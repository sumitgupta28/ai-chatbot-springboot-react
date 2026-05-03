"""
Generate sample_products.xlsx with 100 products across 10 categories.
Each product has a deterministic Picsum Photos image URL.

Usage:
    pip install openpyxl
    python tools/generate_products.py
Output: sample_products.xlsx in the project root.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import Workbook

PRODUCTS = [
    # Electronics (E001–E010)
    ("E001", "Sony WH-1000XM5 Wireless Noise-Cancelling Headphones", "Electronics", "Sony",
     "Industry-leading noise cancellation with up to 30-hour battery life. Multipoint connection for two devices simultaneously.", 349.99, 4.8, 245),
    ("E002", "Apple AirPods Pro (2nd Generation)", "Electronics", "Apple",
     "Active noise cancellation, transparency mode, and adaptive audio. MagSafe charging case with USB-C.", 249.00, 4.7, 512),
    ("E003", "Samsung 65-inch QLED 4K Smart TV", "Electronics", "Samsung",
     "Quantum Dot technology with 4K resolution, HDR10+, and built-in Alexa. Sleek bezel-less design.", 1299.99, 4.6, 87),
    ("E004", "Dell XPS 15 Laptop Intel Core i7", "Electronics", "Dell",
     "15.6-inch OLED display, 16GB RAM, 512GB SSD. Thin and light design with all-day battery life.", 1799.00, 4.5, 63),
    ("E005", "Logitech MX Master 3S Wireless Mouse", "Electronics", "Logitech",
     "8K DPI sensor, quiet clicks, ultra-fast MagSpeed scrolling. Works on any surface including glass.", 99.99, 4.9, 320),
    ("E006", "iPad Air 5th Generation 256GB", "Electronics", "Apple",
     "M1 chip, 10.9-inch Liquid Retina display, Touch ID. Compatible with Apple Pencil 2nd generation.", 749.00, 4.8, 198),
    ("E007", "GoPro HERO12 Black Action Camera", "Electronics", "GoPro",
     "5.3K video, HyperSmooth 6.0 stabilization, waterproof to 33ft. HDR video and photo capability.", 399.99, 4.6, 142),
    ("E008", "Anker 200W USB-C Charging Station", "Electronics", "Anker",
     "7-port charging hub with 200W total output. Charge laptops, tablets, and phones simultaneously.", 89.99, 4.7, 415),
    ("E009", "Ring Video Doorbell Pro 2", "Electronics", "Ring",
     "Head-to-toe HD+ video, 3D motion detection, bird's-eye view. Works with Alexa. Wired installation.", 249.99, 4.4, 276),
    ("E010", "Bose SoundLink Flex Bluetooth Speaker", "Electronics", "Bose",
     "Waterproof and dustproof portable speaker with 12-hour battery. PositionIQ technology for best sound.", 149.00, 4.7, 389),

    # Clothing (C001–C010)
    ("C001", "Levi's 511 Slim Fit Jeans", "Clothing", "Levi's",
     "Classic slim fit jeans in stretch denim. Sits below waist with slim through thigh and leg.", 59.99, 4.5, 634),
    ("C002", "Nike Air Force 1 '07 Sneakers White", "Clothing", "Nike",
     "Iconic basketball-inspired sneakers with leather upper and Air cushioning. Timeless all-white colorway.", 110.00, 4.8, 892),
    ("C003", "The North Face Thermoball Eco Jacket", "Clothing", "The North Face",
     "Lightweight packable jacket filled with ThermoBall Eco insulation made from 100% recycled materials.", 199.00, 4.6, 213),
    ("C004", "Patagonia Better Sweater Fleece Jacket", "Clothing", "Patagonia",
     "Classic fleece made from 100% recycled polyester. Warm, comfortable, and Fair Trade certified.", 139.00, 4.9, 167),
    ("C005", "Adidas Ultraboost 23 Running Shoes", "Clothing", "Adidas",
     "Responsive Boost midsole, Primeknit upper, and Continental rubber outsole. Engineered for long-distance running.", 190.00, 4.7, 445),
    ("C006", "Ralph Lauren Classic Fit Oxford Shirt", "Clothing", "Ralph Lauren",
     "100% cotton oxford shirt with button-down collar. Available in multiple colors, perfect for any occasion.", 98.50, 4.5, 523),
    ("C007", "Lululemon Align High-Rise Yoga Pants", "Clothing", "Lululemon",
     "Ultra-soft Nulu fabric with four-way stretch. Naked sensation with no waistband pressure.", 118.00, 4.8, 731),
    ("C008", "Columbia Newton Ridge Hiking Boots", "Clothing", "Columbia",
     "Waterproof leather and mesh upper with Omni-Grip non-marking traction rubber outsole.", 109.99, 4.6, 289),
    ("C009", "Carhartt Insulated Duck Detroit Jacket", "Clothing", "Carhartt",
     "12-ounce 100% ring-spun cotton duck with firm-hand finish. 3M Thinsulate insulation for warmth.", 129.99, 4.7, 178),
    ("C010", "Champion Reverse Weave Hoodie", "Clothing", "Champion",
     "Classic reverse weave sweatshirt with ribbed side panels. Pre-shrunk fleece that maintains shape.", 65.00, 4.6, 892),

    # Books (B001–B010)
    ("B001", "Atomic Habits by James Clear", "Books", "Penguin Random House",
     "An easy and proven way to build good habits and break bad ones. #1 New York Times bestseller.", 18.99, 4.9, 1245),
    ("B002", "The Psychology of Money by Morgan Housel", "Books", "Harriman House",
     "Timeless lessons on wealth, greed, and happiness. 19 short stories exploring the ways people think about money.", 16.99, 4.8, 987),
    ("B003", "Sapiens: A Brief History of Humankind", "Books", "Harper Perennial",
     "How did Homo sapiens come to dominate the planet? A provocative exploration of human history.", 17.99, 4.7, 1102),
    ("B004", "The Lean Startup by Eric Ries", "Books", "Crown Business",
     "How today's entrepreneurs use continuous innovation to create radically successful businesses.", 19.99, 4.6, 756),
    ("B005", "Deep Work by Cal Newport", "Books", "Grand Central Publishing",
     "Rules for focused success in a distracted world. How to produce at an elite level in our information economy.", 15.99, 4.7, 834),
    ("B006", "Thinking, Fast and Slow by Daniel Kahneman", "Books", "Farrar Straus & Giroux",
     "Two systems that drive the way we think. Engaging, startling, and accessible — a major intellectual event.", 17.00, 4.8, 1023),
    ("B007", "The 4-Hour Work Week by Tim Ferriss", "Books", "Crown Archetype",
     "Escape the 9-5, live anywhere, and join the new rich. Practical guide to lifestyle design.", 14.99, 4.4, 678),
    ("B008", "Clean Code by Robert C. Martin", "Books", "Prentice Hall",
     "A handbook of agile software craftsmanship. Best practices for writing readable, maintainable code.", 44.99, 4.7, 543),
    ("B009", "Dune by Frank Herbert", "Books", "Ace Books",
     "Set in the distant future, Dune is a landmark science fiction novel about politics, religion, and ecology.", 16.00, 4.9, 1456),
    ("B010", "The Midnight Library by Matt Haig", "Books", "Viking",
     "Between life and death there is a library. A dazzling, life-affirming novel about all the lives we almost lived.", 15.99, 4.6, 923),

    # Home & Garden (H001–H010)
    ("H001", "Instant Pot Duo 7-in-1 Electric Pressure Cooker", "Home & Garden", "Instant Pot",
     "Pressure cooker, slow cooker, rice cooker, steamer, sauté pan, yogurt maker, and warmer in one.", 99.95, 4.8, 2341),
    ("H002", "Dyson V15 Detect Cordless Vacuum", "Home & Garden", "Dyson",
     "Laser dust detection, powerful suction, HEPA filtration. Up to 60 minutes runtime on single charge.", 699.99, 4.6, 312),
    ("H003", "Weber Spirit II E-310 Gas Grill", "Electronics", "Weber",
     "3-burner propane grill with 529 sq in cooking area. GS4 grilling system with infinity ignition.", 499.00, 4.7, 187),
    ("H004", "Cuisinart 14-Cup Food Processor", "Home & Garden", "Cuisinart",
     "720-watt motor with stainless steel blade. Wide-mouth feed tube, reversible shredding disc.", 199.95, 4.5, 423),
    ("H005", "Philips Hue White and Color Smart Bulbs 4-Pack", "Home & Garden", "Philips",
     "16 million colors and shades of white light. Control via app or voice. Works with Alexa, Google, Apple HomeKit.", 139.99, 4.7, 678),
    ("H006", "LEVOIT Air Purifier Large Room", "Home & Garden", "LEVOIT",
     "Covers 1095 sq ft, H13 True HEPA filter, removes 99.97% of particles. Ultra-quiet sleep mode.", 149.99, 4.8, 1023),
    ("H007", "Casper Sleep Original Foam Mattress Queen", "Home & Garden", "Casper",
     "Zoned Support foam for head-to-toe comfort. Breathable, durable, and pressure-relieving.", 895.00, 4.5, 156),
    ("H008", "Nespresso Vertuo Plus Coffee Maker", "Home & Garden", "Nespresso",
     "Centrifusion extraction technology, 5 cup sizes from espresso to alto. 40-second heat-up time.", 179.00, 4.7, 534),
    ("H009", "iRobot Roomba j7+ Self-Emptying Robot Vacuum", "Home & Garden", "iRobot",
     "P.O.O.P. technology avoids pet waste. Clean Base automatic dirt disposal. Smart mapping with 10x power-lifting suction.", 799.99, 4.5, 234),
    ("H010", "KitchenAid Artisan Stand Mixer 5Qt", "Home & Garden", "KitchenAid",
     "67-point planetary mixing action, 10 speed settings, includes flat beater, dough hook, and wire whip.", 449.99, 4.9, 389),

    # Sports (S001–S010)
    ("S001", "Peloton Bike+ Indoor Exercise Bike", "Sports", "Peloton",
     "23.8-inch rotating HD touchscreen, auto-follow resistance, immersive sound system. Monthly membership required.", 2495.00, 4.4, 89),
    ("S002", "Hydro Flask 32oz Wide Mouth Water Bottle", "Sports", "Hydro Flask",
     "TempShield double-wall vacuum insulation keeps drinks cold 24hrs, hot 12hrs. BPA-free stainless steel.", 44.95, 4.9, 1567),
    ("S003", "Garmin Forerunner 255 GPS Running Watch", "Sports", "Garmin",
     "Training readiness feature, daily suggested workouts, race predictor. Up to 14-day battery life.", 349.99, 4.7, 312),
    ("S004", "Manduka PRO Yoga Mat 6mm", "Sports", "Manduka",
     "High-density cushioning for unmatched support. Non-slip surface, closed-cell technology. Lifetime guarantee.", 120.00, 4.8, 445),
    ("S005", "TRX All-in-One Suspension Training System", "Sports", "TRX",
     "Professional-grade suspension trainer with door anchor, getting started guide. Works at home or gym.", 149.95, 4.7, 567),
    ("S006", "Wilson Evolution Indoor Basketball", "Sports", "Wilson",
     "Composite leather cover with laid-in channel. The #1 selling indoor ball in the US.", 69.99, 4.8, 892),
    ("S007", "Osprey Atmos AG 65L Hiking Backpack", "Sports", "Osprey",
     "Anti-gravity suspension system, hipbelt pockets, integrated raincover. Best-in-class fit and comfort.", 290.00, 4.9, 234),
    ("S008", "Fitbit Charge 6 Fitness Tracker", "Sports", "Fitbit",
     "Built-in GPS, heart rate monitoring, Active Zone Minutes. Up to 7 days battery, 40+ exercise modes.", 159.95, 4.5, 678),
    ("S009", "Callaway Strata 16-Piece Golf Set", "Sports", "Callaway",
     "Complete set including driver, fairway wood, hybrids, irons, wedge, putter, and stand bag.", 299.99, 4.6, 156),
    ("S010", "Bowflex SelectTech 552 Adjustable Dumbbells Pair", "Sports", "Bowflex",
     "Adjusts from 5 to 52.5 lbs. Replaces 15 sets of weights. Easy turn dial for quick weight change.", 549.00, 4.7, 289),

    # Beauty (BT001–BT010)
    ("BT001", "CeraVe Moisturizing Cream 19oz", "Beauty", "CeraVe",
     "Rich, non-greasy moisturizer with ceramides and hyaluronic acid. Fragrance-free, non-comedogenic.", 19.99, 4.8, 2345),
    ("BT002", "Dyson Airwrap Multi-Styler Complete", "Beauty", "Dyson",
     "Styles and dries simultaneously with no extreme heat. Coanda effect to attract and wrap hair.", 599.99, 4.6, 312),
    ("BT003", "Charlotte Tilbury Pillow Talk Lip Liner", "Beauty", "Charlotte Tilbury",
     "Iconic dusky-rose pink nude shade. Waterproof formula that lasts all day. The world's No.1 lip liner.", 26.00, 4.9, 1234),
    ("BT004", "The Ordinary Hyaluronic Acid 2% + B5", "Beauty", "The Ordinary",
     "Multi-depth hyaluronic acid serum with vitamin B5. Hydrates skin at multiple skin-surface layers.", 7.90, 4.7, 3456),
    ("BT005", "Olaplex No. 3 Hair Perfector 3.3oz", "Beauty", "Olaplex",
     "Reduces breakage and visibly strengthens hair. Use weekly as a pre-shampoo treatment.", 30.00, 4.8, 1678),
    ("BT006", "Tatcha The Water Cream Moisturizer", "Beauty", "Tatcha",
     "Oil-free anti-aging water cream with Japanese wild rose and leopard lily. Weightless hydration.", 68.00, 4.7, 456),
    ("BT007", "Urban Decay All Nighter Setting Spray", "Beauty", "Urban Decay",
     "Sets makeup for up to 16 hours. Temperature control technology. Lightweight micro-fine mist.", 34.00, 4.6, 1234),
    ("BT008", "Necessaire The Body Serum", "Beauty", "Necessaire",
     "Fragrance-free niacinamide, hyaluronic acid, and vitamin C body serum for all skin types.", 45.00, 4.7, 567),
    ("BT009", "Foreo Luna 4 Facial Cleansing Device", "Beauty", "Foreo",
     "T-Sonic pulsations remove 99.5% of makeup and impurities. 16 intensity levels, USB rechargeable.", 199.00, 4.5, 234),
    ("BT010", "Fenty Beauty Pro Filt'r Soft Matte Foundation", "Beauty", "Fenty Beauty",
     "50 shades, long-wearing 24-hour foundation. Oil-free, buildable medium-to-full coverage.", 38.00, 4.7, 892),

    # Toys (T001–T010)
    ("T001", "LEGO Creator Expert Eiffel Tower 10307", "Toys", "LEGO",
     "10,001 pieces, 1.5m tall when complete. Intricate lattice structure with observation decks.", 629.99, 4.9, 145),
    ("T002", "Nintendo Switch OLED Model", "Toys", "Nintendo",
     "7-inch OLED screen, enhanced audio, 64GB storage. Play at home or on the go in three modes.", 349.99, 4.8, 512),
    ("T003", "Barbie Dreamhouse Playset", "Toys", "Mattel",
     "3-story, 8-room dollhouse with pool, slide, and elevator. Over 75 accessories included.", 199.99, 4.7, 234),
    ("T004", "Hot Wheels Ultimate Garage Playset", "Toys", "Hot Wheels",
     "5-level parking garage with car wash, chomper, and helicopter landing pad. Holds 140+ cars.", 99.99, 4.6, 345),
    ("T005", "Magna-Tiles Clear Colors 100-Piece Set", "Toys", "Magna-Tiles",
     "Magnetic building tiles encourage STEM learning and creativity. Compatible with all Magna-Tiles sets.", 139.99, 4.9, 678),
    ("T006", "Melissa & Doug Wooden Building Blocks 100-Piece", "Toys", "Melissa & Doug",
     "Solid hardwood blocks in 4 colors and 9 shapes. Promotes spatial reasoning and creativity.", 29.99, 4.8, 1023),
    ("T007", "DJI Mini 3 Pro Drone with RC Remote", "Toys", "DJI",
     "Under 249g, 4K/60fps video, 48MP photos, obstacle sensing. 34-min max flight time.", 759.00, 4.7, 167),
    ("T008", "Exploding Kittens Card Game", "Toys", "Exploding Kittens",
     "A card game for people who are into kittens and explosions. For 2-5 players, ages 7+.", 19.99, 4.7, 1567),
    ("T009", "Sphero Mini App-Controlled Robot Ball", "Toys", "Sphero",
     "Golf ball-sized robotic ball with colorful LED lights. Drive, play games, or learn to code.", 49.99, 4.5, 423),
    ("T010", "National Geographic Break Open 10 Premium Geodes Kit", "Toys", "National Geographic",
     "10 whole geodes to crack open with safety goggles and learning guide. Real crystals inside.", 29.99, 4.8, 892),

    # Food (F001–F010)
    ("F001", "Death Wish Coffee Whole Bean 1lb", "Food", "Death Wish Coffee",
     "The world's strongest coffee. USDA certified organic, fair trade. Bold, smooth, with cherry and chocolate notes.", 19.99, 4.7, 1234),
    ("F002", "Ghirardelli Premium Hot Cocoa Variety Pack", "Food", "Ghirardelli",
     "6 canisters of premium hot chocolate including Double Chocolate, Peppermint Mocha, and Salted Caramel.", 39.99, 4.8, 678),
    ("F003", "Justin's Classic Almond Butter Squeeze Pack 10ct", "Food", "Justin's",
     "USDA organic, gluten-free squeeze packs. No palm oil, just dry roasted almonds and palm fruit oil.", 14.99, 4.7, 892),
    ("F004", "RXBAR Protein Bar Variety Pack 24ct", "Food", "RXBAR",
     "Real food protein bars with 12g protein and no BS. Whole dates, egg whites, and almonds.", 38.99, 4.6, 1456),
    ("F005", "Bob's Red Mill Organic Rolled Oats 32oz", "Food", "Bob's Red Mill",
     "100% whole grain, USDA organic rolled oats. No added salt, sugar, or oil. High in fiber.", 7.99, 4.8, 2345),
    ("F006", "Graza Sizzle Extra Virgin Olive Oil 750ml", "Food", "Graza",
     "Single-origin picual olives from Jaén, Spain. Fruity, buttery taste. For everyday cooking.", 21.00, 4.8, 1023),
    ("F007", "Tony's Chocolonely Milk Chocolate Bar 6-Pack", "Food", "Tony's Chocolonely",
     "32% milk chocolate bars made with 100% slave-free cocoa. Unequally divided as a statement.", 29.99, 4.8, 789),
    ("F008", "Kind Dark Chocolate Nuts & Sea Salt Bar 24ct", "Food", "Kind",
     "6g protein, 5g fiber, low glycemic, gluten-free, non-GMO. Dark chocolate coated almond and peanut bar.", 32.99, 4.7, 1678),
    ("F009", "Chomps Original Beef Jerky Snack Sticks 24ct", "Food", "Chomps",
     "Keto, paleo, Whole30 approved. Grass-fed & finished beef with sea salt and pepper seasoning.", 41.99, 4.6, 934),
    ("F010", "Jacobsen Salt Co. Pure Kosher Sea Salt 28oz", "Food", "Jacobsen Salt Co.",
     "Hand-harvested from the Oregon coast. Clean, bright flavor perfect for finishing and cooking.", 24.00, 4.9, 567),

    # Automotive (A001–A010)
    ("A001", "Anker Roav Dash Cam C2 Pro", "Automotive", "Anker",
     "1080P Full HD recording with night vision, wide 140° view angle, GPS tracking, and Wi-Fi.", 79.99, 4.5, 456),
    ("A002", "Chemical Guys HOL126 14-Item Car Wash Kit", "Automotive", "Chemical Guys",
     "Complete detailing kit with snow foam, car wash soap, microfiber towels, and wheel cleaner.", 84.99, 4.7, 678),
    ("A003", "NOCO Boost Plus GB40 1000A Jump Starter", "Automotive", "NOCO",
     "Portable lithium jump starter for up to 6-liter gas and 3-liter diesel engines. USB charging port.", 99.95, 4.8, 1234),
    ("A004", "WeatherTech FloorLiners Front and Rear Set", "Automotive", "WeatherTech",
     "Custom-fit laser-measured floor mats. High-density tri-extruded material, vehicle-specific design.", 149.95, 4.8, 892),
    ("A005", "Garmin DriveSmart 76 GPS Navigator 7-inch", "Automotive", "Garmin",
     "7-inch display with bright sunlight visibility, live traffic and weather, and voice-activated navigation.", 199.99, 4.5, 312),
    ("A006", "Armor All Car Interior Cleaner Kit 8-Piece", "Automotive", "Armor All",
     "Complete cleaning set for dashboard, seats, glass, tires, and exterior. UV protection included.", 24.99, 4.6, 1567),
    ("A007", "Thule Pulse Alpine Rooftop Cargo Box 16cf", "Automotive", "Thule",
     "16 cubic feet of cargo space, dual-side opening, low-profile aerodynamic design. Universal fit system.", 499.95, 4.7, 134),
    ("A008", "BESTEK 300W Power Inverter for Car", "Automotive", "BESTEK",
     "DC 12V to 110V AC inverter with dual USB charging ports. Ideal for road trips and emergencies.", 35.99, 4.6, 2345),
    ("A009", "Michelin FastFix Tire Inflation Kit", "Automotive", "Michelin",
     "Portable digital tire inflator with LED light. Auto-shut off when target pressure reached.", 49.99, 4.7, 678),
    ("A010", "Covercraft Custom Fit Car Cover", "Automotive", "Covercraft",
     "Vehicle-specific pattern for perfect fit. Weather Block fabric resists moisture, UV, and dust.", 189.00, 4.5, 234),

    # Office (O001–O010)
    ("O001", "Ergotron LX Dual Monitor Arm", "Office", "Ergotron",
     "Holds two displays up to 32 inches and 20 lbs each. Smooth articulating arm with full range of motion.", 239.99, 4.8, 567),
    ("O002", "Herman Miller Aeron Office Chair Size B", "Office", "Herman Miller",
     "Iconic ergonomic chair with PostureFit SL back support, 8Z Pellicle mesh, and adjustable arms.", 1495.00, 4.9, 89),
    ("O003", "Autonomous SmartDesk Pro Standing Desk", "Office", "Autonomous",
     "Height adjustable from 26.2-52 inches. Dual motor, 4 memory presets, 300 lb capacity.", 699.00, 4.6, 234),
    ("O004", "Moleskine Classic Notebook Hardcover Large Ruled", "Office", "Moleskine",
     "240 pages, acid-free paper with rounded corners, ribbon bookmark, and elastic closure.", 22.95, 4.7, 3456),
    ("O005", "Brother HL-L2395DW Wireless Laser Printer", "Office", "Brother",
     "Wireless monochrome laser printer with duplex printing. Up to 36ppm, 250-sheet tray.", 219.99, 4.6, 456),
    ("O006", "Pilot G2 Premium Retractable Pens 20-Pack", "Office", "Pilot",
     "Smooth gel ink with comfortable rubber grip. Retractable tip in assorted colors.", 19.99, 4.8, 5678),
    ("O007", "Fellowes Powershred 79Ci Cross-Cut Shredder", "Office", "Fellowes",
     "Shreds 16 sheets per pass into cross-cut particles. Jam-proof technology, 100% jam-free guarantee.", 179.99, 4.6, 312),
    ("O008", "Elfa Freestanding Desk with Drawer Unit", "Office", "Elfa",
     "Customizable freestanding desk system with drawer unit. Adjustable shelves and cable management.", 449.00, 4.5, 145),
    ("O009", "BenQ ScreenBar Plus Monitor Light with Remote", "Office", "BenQ",
     "Auto-dimming monitor lamp with wireless dial remote. No screen glare, 500+ lux illumination.", 109.00, 4.8, 678),
    ("O010", "Staples Hyken Mesh Task Chair Black", "Office", "Staples",
     "Ergonomic mesh back with breathable design, adjustable lumbar support, and multi-tilt mechanism.", 299.99, 4.5, 423),
]

def generate_image_url(product_id):
    return f"https://picsum.photos/seed/{product_id}/400/300"

def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    headers = ["ProductID", "Name", "Category", "Brand", "Description", "Price", "ImageUrl", "Rating", "StockCount"]

    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 25

    for row_num, product in enumerate(PRODUCTS, 2):
        pid, name, category, brand, description, price, rating, stock = product
        image_url = generate_image_url(pid)
        ws.cell(row=row_num, column=1, value=pid)
        ws.cell(row=row_num, column=2, value=name)
        ws.cell(row=row_num, column=3, value=category)
        ws.cell(row=row_num, column=4, value=brand)
        ws.cell(row=row_num, column=5, value=description)
        ws.cell(row=row_num, column=6, value=price)
        ws.cell(row=row_num, column=7, value=image_url)
        ws.cell(row=row_num, column=8, value=rating)
        ws.cell(row=row_num, column=9, value=stock)

    col_widths = [10, 50, 18, 20, 70, 10, 50, 8, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    ws.freeze_panes = "A2"

    output_path = "sample_products.xlsx"
    wb.save(output_path)
    print(f"Generated {len(PRODUCTS)} products → {output_path}")

if __name__ == "__main__":
    main()
